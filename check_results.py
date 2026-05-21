import openpyxl

# Открываем наш эксель файл
wb = openpyxl.load_workbook("avito_results.xlsx")
ws = wb.active

print("\n🎯 ТОП ПРЕДЛОЖЕНИЙ ИЗ EXCEL (Снайперский режим):")
print("-" * 130)

# Читаем данные начиная с 3-й строки
for row in range(3, ws.max_row + 1):
    title = ws.cell(row=row, column=3).value
    
    # Пропускаем пустые строки или итоговую строчку внизу
    if not title or "Всего" in str(title): 
        continue
        
    # Безопасное извлечение данных (если пусто — ставим дефолтное значение)
    score = ws.cell(row=row, column=2).value
    score = int(score) if score is not None else 0
    
    price = ws.cell(row=row, column=4).value
    price = int(price) if price is not None else 0
    
    city = ws.cell(row=row, column=5).value or "Не указан"
    kit = ws.cell(row=row, column=8).value or "—"
    
    # Выводим красиво в консоль
    print(f"⭐ Оценка: {score:2} | 💰 {price:6,d} руб. | 🎁 {kit:10s} | 📍 {str(city)[:15]:15s} | 📱 {str(title)[:35]}")

print("-" * 130)