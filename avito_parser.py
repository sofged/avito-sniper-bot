import asyncio
import re
import json
import os
from urllib.parse import urlencode
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================
# ⚙️ ОСНОВНЫЕ НАСТРОЙКИ
# ==========================================
PRICE_MIN = 7000          
PRICE_MAX = 15000         
MAX_PAGES  = 50           
OUTPUT_FILE = "avito_results.xlsx" 
USER_CITY = "ростов-на-дону"

# Конфигурация кэша
CACHE_FILE = "avito_cache.json"

def load_cache() -> dict:
    """Load cached ad details from CACHE_FILE if it exists."""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_cache(cache: dict) -> None:
    """Persist the cache dictionary to CACHE_FILE."""
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)       

SEARCH_QUERIES = [
    "motorola",
    "nothing",
    "google pixel",
    "oneplus",
]
BASE_URL = "https://www.avito.ru/rossiya/telefony/smartfony-ASgBAgICAUSSA8YQ"

def make_url(query: str, page: int = 1) -> str:
    """Генерирует ссылку для поиска на Авито с фильтрами по цене и доставке."""
    params = {"q": query, "pmin": PRICE_MIN, "pmax": PRICE_MAX, "cd": 1}
    if page > 1: params["p"] = page
    return f"{BASE_URL}?{urlencode(params)}"

async def is_captcha_page(page) -> bool:
    """Проверяет, не заблокировал ли нас Авито капчей (по заголовку страницы)."""
    title = (await page.title()).lower()
    return "ограничен" in title or "робот" in title or "captcha" in title

async def fetch_description(page, url: str) -> dict:
    """
    Заходит в конкретное объявление и вытягивает характеристики (Состояние, Экран, АКБ, Рейтинг, Год регистрации).
    Использует JavaScript для поиска текста по всей странице, обходя скрытые теги Авито.
    """
    result = {
        "description": "Описание не найдено", "condition": "Не указано", 
        "screen": "Не указано", "battery": "—", "rating": 0.0, "reviews": 0, "reg_year": 0
    }
    
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=15_000)
        await asyncio.sleep(3)
        
        # Если вылезла капча - ждем, пока пользователь ее решит руками
        if await is_captcha_page(page):
            print(f"\n⚠ Капча при переходе. Реши в браузере...")
            for _ in range(60):
                await asyncio.sleep(3)
                if not await is_captcha_page(page): break
            else:
                result["description"] = "Капча при загрузке"
                return result

        # Парсим текст описания объявления
        desc_el = await page.query_selector('[data-marker="item-view/item-description"]') or await page.query_selector('[itemprop="description"]')
        if desc_el:
            result["description"] = (await desc_el.inner_text()).strip()

        # Выполняем JS-скрипт внутри страницы для поиска нужных данных в любом месте текста
        js_extractor = r"""
        () => {
            let res = {condition: "Не указано", screen: "Не указано", battery: "—", rating: 0.0, reviews: 0, reg_year: 0};
            let text = document.body.innerText;
            
            // Ищем Состояние и Экран
            let condMatch = text.match(/Состояние:?\s*(Новое|Отличное|Хорошее|Удовлетворительное|Требуется ремонт|На запчасти)/i);
            if (condMatch) res.condition = condMatch[1].charAt(0).toUpperCase() + condMatch[1].slice(1).toLowerCase();
            
            let screenMatch = text.match(/Экран:?\s*(Без дефектов|1[-–]2 мелкие царапины|Много мелких царапин|Глубокие царапины|Трещины|Засветы|Выгорание|Полосы и битые пиксели|Не работает)/i);
            if (screenMatch) res.screen = screenMatch[1].charAt(0).toUpperCase() + screenMatch[1].slice(1).toLowerCase().replace('1-2', '1–2');
            
            // Ищем процент АКБ
            let batMatch = text.match(/(?:акб|батарея|емкость|состояние аккумулятора)[\s:-]*(\d{2,3})\s*%/i);
            if (batMatch) res.battery = batMatch[1] + "%";

            // Ищем рейтинг и отзывы
            let ratingMatch = text.match(/(\d[,.]\d)\s*(?:★\s*)?\n?\s*(\d+)\s+отзыв/i);
            if (ratingMatch) {
                res.rating = parseFloat(ratingMatch[1].replace(',', '.'));
                res.reviews = parseInt(ratingMatch[2]);
            }
            
            // Ищем дату регистрации аккаунта (например: "На Авито с апреля 2023")
            let regMatch = text.match(/На Авито с [а-яА-Я]+ (\d{4})/i);
            if (regMatch) res.reg_year = parseInt(regMatch[1]);

            return res;
        }
        """
        data = await page.evaluate(js_extractor)
        if data: result.update(data)

    except Exception as e:
        result["description"] = f"Ошибка: {e}"
    
    return result


async def parse_page(page, url: str, context, page_num: int) -> list[dict]:
    """Сканирует общую страницу поиска, собирает карточки товаров и делает первичный фильтр."""
    try:
        # Смягчаем условия загрузки: domcontentloaded быстрее, чем load
        await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        
        # Пытаемся дождаться networkidle, но не падаем, если оно не случилось
        try:
            await page.wait_for_load_state("networkidle", timeout=10_000)
        except Exception:
            pass # Игнорируем таймаут ожидания сети, если DOM уже готов
            
        await asyncio.sleep(2)
        
        # Проверка: не перебросило ли нас на другую страницу (например, на первую страницу без p=)
        current_url = page.url
        if f"p={page_num}" not in current_url and page_num > 1:
            print(f"  ⚠ Авито сбросил пагинацию (перенаправил на {current_url})")
            return None
    except Exception as e:
        print(f"  ⚠ Ошибка при загрузке {url}: {e}")
        # Если загрузка совсем упала, возвращаем пустой список
        return []

    if await is_captcha_page(page):
        print("  ⚠ Обнаружена капча!")
        return []

    try:
        # Проверяем наличие карточек. Если они есть — работаем, даже если страница не «догрузилась» до конца
        await page.wait_for_selector('[data-marker="item"]', timeout=10_000)
    except Exception:
        # Если карточек нет совсем — значит либо конец, либо блок
        return []

    cards = await page.query_selector_all('[data-marker="item"]')
    items = []

    for card in cards:
        try:
            # Извлекаем базовые данные с карточки (название, цена, город, ссылка)
            title_el = await card.query_selector('[itemprop="name"]') or await card.query_selector('[data-marker="item-title"]')
            if not title_el: continue
            title = (await title_el.inner_text()).strip()

            price_el = await card.query_selector('[itemprop="price"]')
            price = int(await price_el.get_attribute("content") or 0) if price_el else 0

            link_el = await card.query_selector('[data-marker="item-title"]')
            href = await link_el.get_attribute("href") if link_el else ""
            url_item = f"https://www.avito.ru{href}" if href and href.startswith("/") else href

            geo_el = await card.query_selector('[data-marker="item-address"]')
            city = ""
            if geo_el:
                try:
                    city_text = await geo_el.inner_text()
                    city = city_text.strip().split(",")[0].strip()
                except Exception:
                    pass

            # --- ПЕРВИЧНЫЙ ФИЛЬТР ---
            STOP_WORDS = [
                "чехол", "коробка", "аккумулятор", "акб", "дисплей", "экран", "стекло", "камера",
                "плата", "запчасти", "на запчасти", "шлейф", "корпус", "динамик", "зарядка", 
                "ремонт", "наушники", "watch", "buds", "часы", "разбит", "трещина", "выгорание",
                "кейс", "бампер", "кулер", "защитное", "кабель", "провод", "блок питания"
            ]
            
            title_lower = title.lower()
            if any(w in title_lower for w in STOP_WORDS): continue

            # Строгий фильтр по нужным моделям телефонов
            is_valid_model = False
            
            # 1. Nothing Phone
            if "nothing" in title_lower or "phone" in title_lower:
                if "cmf" in title_lower or "lite" in title_lower: pass 
                # Убрали 2a и 2 pro, оставили 2, 3, 4 и их вариации
                elif re.search(r'(nothing|phone)\s*\(?(2|3|4)\b', title_lower):
                    if not re.search(r'2\s*a|2\s*pro|lite', title_lower):
                        is_valid_model = True
            
            # 2. Google Pixel
            elif "pixel" in title_lower or "пиксель" in title_lower:
                if re.search(r'(pixel|пиксель)\s*(8|9|9\s*pro|9pro|9\s*a|9a)\b', title_lower):
                    is_valid_model = True
            
            # 3. Motorola
            elif any(brand in title_lower for brand in ["motorola", "moto", "моторола"]):
                allowed_moto = [
                    r"edge\s*40", r"edge\s*50", r"edge\s*60", r"edge\s*70",
                    r"\bx\s*40\b", r"\bs\s*50\b", r"\bs\s*60\b"
                ]
                if any(re.search(m, title_lower) for m in allowed_moto):
                    # Исключаем 20, 30 серии, модель S30 и X30
                    if not re.search(r'edge\s*(20|30)|s\s*30|x\s*30', title_lower):
                        is_valid_model = True

            # 4. OnePlus
            elif "oneplus" in title_lower or "ванплас" in title_lower:
                # Ищем 11, 11R, 12, 12R, Nord 3/4/5, Ace 2/3/5 и их Pro/V версии
                # Используем более гибкий поиск для Ace 2/3/5, чтобы ловить 2V, 3V, 5V, 2 Pro и т.д.
                op_regex = r'(oneplus|ванплас)\s+(11\s*r?|12\s*r?|nord\s*[345]|ace\s*[235][rv\s]*|ace\s*[235]\s*pro)\b'
                if re.search(op_regex, title_lower):
                    # ИСКЛЮЧАЕМ Nord CE 3 и Lite версии
                    if not re.search(r'ce\s*3|lite', title_lower):
                        is_valid_model = True
            
            if not is_valid_model: continue

            if price and (PRICE_MIN <= price <= PRICE_MAX):
                items.append({"title": title, "price": price, "city": city, "url": url_item})

        except Exception:
            continue

    return items


def save_to_excel(data: list[dict], filename: str):
    """Сохраняет результаты в Excel с раскраской ячеек и автофильтрами."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Авито — ТОП Смартфоны"

    DARK, ACCENT, LIGHT_ROW, WHITE, GREEN, TEXT_DARK, GOLD, RED = "111827", "6366F1", "F9FAFB", "FFFFFF", "D1FAE5", "1F2937", "F59E0B", "FEE2E2"
    border = Border(left=Side(style="thin", color="E5E7EB"), right=Side(style="thin", color="E5E7EB"), 
                    top=Side(style="thin", color="E5E7EB"), bottom=Side(style="thin", color="E5E7EB"))

    # Шапка
    cols = ["№", "Оценка", "Название", "Цена (₽)", "Город", "Рейтинг", "АКБ", "Комплект", "Состояние", "Экран", "Описание", "Ссылка"]
    ws.merge_cells(f"A1:{chr(64+len(cols))}1")
    header_main = ws["A1"]
    header_main.value = f"🎯 СНАЙПЕРСКАЯ ПОДБОРКА │ {PRICE_MIN:,} – {PRICE_MAX:,} ₽"
    header_main.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    header_main.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    header_main.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Названия колонок
    col_widths = [5, 10, 35, 12, 18, 22, 10, 15, 15, 22, 60, 25]
    for i, (col_name, width) in enumerate(zip(cols, col_widths), 1):
        cell = ws.cell(row=2, column=i, value=col_name)
        cell.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[2].height = 22

    # Заполнение строк
    row_num = 3
    for idx, item in enumerate(data, 1):
        desc = item.get("description", "—")
        if len(desc) > 300: desc = desc[:297] + "..." 
        
        # --- ФОРМАТИРОВАНИЕ РЕЙТИНГА И ДАТЫ РЕГИСТРАЦИИ ---
        rating_text = f"{item.get('rating', 0.0)} ({item.get('reviews', 0)} отз.)"
        rating_color = None
        
        reviews = item.get('reviews', 0)
        reg_year = item.get('reg_year', 0)
        
        if reviews == 0:
            if reg_year == 0:
                rating_text = "⚠ Нет отзывов"
            elif reg_year >= 2024:
                rating_text = f"⚠ Опасно (Свежий акк {reg_year})"
                rating_color = RED
            else:
                rating_text = f"Без отзывов (С {reg_year} г.)"
                rating_color = "FEF08A"

        row_data = [idx, item.get("score", 0), item.get("title", ""), item.get("price", 0), item.get("city", ""), rating_text, 
                    item.get("battery", "—"), item.get("kit", "❓ Уточнить"), item.get("condition", "Не указано"), 
                    item.get("screen", "Не указано"), desc, item.get("url", "")]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = Font(name="Calibri", size=10, color=TEXT_DARK)
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [3, 11]))
            cell.border = border
            if idx % 2 == 0: cell.fill = PatternFill(start_color=LIGHT_ROW, end_color=LIGHT_ROW, fill_type="solid")

        # Раскраска Оценки
        score = item.get("score", 0)
        score_cell = ws.cell(row=row_num, column=2)
        score_cell.font = Font(bold=True, size=12, color="FFFFFF" if score >= 5 else TEXT_DARK)
        score_cell.fill = PatternFill(start_color="10B981" if score >= 5 else "D1D5DB", fill_type="solid")
        score_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=row_num, column=4).font = Font(bold=True, color="065F46")
        ws.cell(row=row_num, column=4).number_format = '#,##0 ₽'
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center", vertical="center") 
        
        if item.get("kit") == "✅ Полный":
            ws.cell(row=row_num, column=8).font = Font(color="059669", bold=True)
            
        # Раскраска рейтинга
        if rating_color:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color=rating_color, end_color=rating_color, fill_type="solid")
            if rating_color == RED: ws.cell(row=row_num, column=6).font = Font(color="B91C1C", bold=True)

        url_cell = ws.cell(row=row_num, column=12)
        url_cell.hyperlink = item.get("url", "")
        url_cell.font = Font(color="4F46E5", underline="single")
        url_cell.value = "Открыть"

        ws.row_dimensions[row_num].height = 45  
        row_num += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{chr(64+len(cols))}{row_num-1}"
    wb.save(filename)
    return len(data)


# ==========================================
# 🚀 ГЛАВНЫЙ АЛГОРИТМ ПАРСЕРА
# ==========================================
async def main():
    try:
        all_items = []
        print("🚀 Запуск парсера: СНАЙПЕРСКИЙ РЕЖИМ (С ОЦЕНКАМИ)")

        # Load cache
        cache = load_cache()
        cached_used = 0
        new_added = 0

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=False, args=["--disable-blink-features=AutomationControlled"])
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                locale="ru-RU", viewport={"width": 1366, "height": 900})
            await context.add_init_script("Object.defineProperty(navigator, 'webdriver', { get: () => undefined });")
            page = await context.new_page()

            # --- ЭТАП 1: СБОР ССЫЛОК СО СТРАНИЦ ПОИСКА ---
            for query in SEARCH_QUERIES:
                print(f"\n🔍 Поиск: «{query}»")
                empty_attempts = 0
                query_seen_ids = set() # Используем ID для детекции дублей
                
                for page_num in range(1, MAX_PAGES + 1):
                    try:
                        url = make_url(query, page_num)
                        print(f"  Страница {page_num}... ", end="", flush=True)
                        
                        items = await parse_page(page, url, context, page_num)
                        
                        # Если функция вернула None — это жесткий стоп (редирект пагинации)
                        if items is None:
                            break
                        
                        # Извлекаем ID из ссылок для точной дедупликации
                        for it in items:
                            match = re.search(r'_(\d+)$', it["url"].split("?")[0])
                            it["id"] = match.group(1) if match else it["url"]

                        new_items = [it for it in items if it["id"] not in query_seen_ids]
                        
                        if items:
                            print(f"Найдено: {len(items)} (новых: {len(new_items)})")
                            for it in new_items:
                                query_seen_ids.add(it["id"])
                            all_items.extend(new_items)
                        else:
                            print("Найдено: 0")

                        # ЛОГИКА ОСТАНОВКИ
                        if not new_items:
                            empty_attempts += 1
                            
                            # Если на странице есть результаты, но они все старые — это петля пагинации (Авито показывает одни и те же VIP-объявления)
                            if items and len(items) > 0:
                                print("  🛑 Повторяющиеся результаты (петля пагинации).")
                                break
                            
                            if empty_attempts >= 20:
                                print("  🛑 Пустые страницы, переходим к следующему запросу.")
                                break
                        else:
                            empty_attempts = 0

                        await asyncio.sleep(4) # Пауза между страницами
                    except Exception as e:
                        print(f"  ⚠ Сбой на странице {page_num}: {e}")
                        continue
            
            # Deduplicate items by URL (на всякий случай, если в разных запросах попались одинаковые ссылки)
            unique_items = []
            seen_global = set()
            for it in all_items:
                if it["url"] not in seen_global:
                    seen_global.add(it["url"])
                    unique_items.append(it)
            
            # --- ЭТАП 2: Анализ каждого объявления ---
            filtered_items = []
            seen_content = set() # Для удаления дублей по тексту описания
            
            if unique_items:
                print(f"📝 Анализ ({len(unique_items)} объявл.): Скрытые дефекты, Доставка, Комплект, Рейтинг...")
                for idx, item in enumerate(unique_items, 1):
                    try:
                        url = item["url"]
                        is_from_cache = False
                        
                        if url in cache:
                            item.update(cache[url])
                            is_from_cache = True
                        else:
                            print(f"  [{idx}/{len(unique_items)}] {item['title'][:20]}... ", end="", flush=True)
                            details = await fetch_description(page, url)
                            item.update(details)
                            await asyncio.sleep(3)
                        
                        # --- ПРОВЕРКА НА ДУБЛИКАТЫ ПО КОНТЕНТУ ---
                        desc_text = item.get("description", "").strip().lower()
                        # Используем очищенный кусок описания для выявления шаблонов магазинов (игнорируя мелкие изменения)
                        clean_desc = re.sub(r'\s+', '', desc_text[:400])
                        
                        if len(clean_desc) > 50 and clean_desc in seen_content:
                            if not is_from_cache: print("✕ (Дубликат описания/Магазин-спам)")
                            continue
                        if len(clean_desc) > 50:
                            seen_content.add(clean_desc)

                        # --- ФИЛЬТРЫ (применяем ко всем, даже к кешу) ---
                        desc_lower = item.get("description", "").lower()
                        title_lower = item.get("title", "").lower()
                        
                        # 1. Повторная проверка стоп-слов в названии
                        STOP_WORDS_EXTRA = ["кейс", "бампер", "кулер", "защитное", "кабель", "провод", "блок питания", "дисплей", "экран"]
                        if any(w in title_lower for w in STOP_WORDS_EXTRA):
                            if not is_from_cache: print("✕ (Аксессуар/Запчасть)")
                            continue

                        # 🛑 Фильтр скрытых проблем
                        stop_issues = [
                            "mdm", "мдм", "soft unlock", "софт анлок", "операторский", "демо", "demo", "att", "verizon",
                            "вскрывался", "ремонтировался", "заменен", "менялся", "не работает отпечаток", "отвал",
                            "мерцает", "заблокирован", "icloud", "гугл аккаунт", "google аккаунт", "пятн", "требуется ремонт"
                        ]
                        if any(bad in desc_lower for bad in stop_issues):
                            if not is_from_cache: print("✕ (Скрытый дефект/Блокировка)")
                            continue

                        if "только обмен" in desc_lower:
                            if not is_from_cache: print("✕ (Только обмен)")
                            continue

                        # 🛑 Фильтр доставки
                        no_delivery = ["без доставки", "доставки нет", "авито доставки нет", "не отправляю", "только личная встреча", "не отправлю", "без пересыла"]
                        if any(word in desc_lower for word in no_delivery):
                            if USER_CITY not in item.get("city", "").lower():
                                if not is_from_cache: print("✕ (Нет доставки, другой город)")
                                continue

                        # 🛑 Фильтр состояния
                        cond_lower = item.get("condition", "Не указано").lower()
                        scr_lower = item.get("screen", "Не указано").lower()
                        if cond_lower != "не указано" and not any(good in cond_lower for good in ["отлично", "хороше", "ново"]):
                            if not is_from_cache: print("✕ (Убитое состояние)")
                            continue
                        if scr_lower != "не указано" and not any(good in scr_lower for good in ["без дефект", "1-2", "1–2", "царапин"]):
                            if not is_from_cache: print("✕ (Разбитый экран)")
                            continue

                        # 🛑 Фильтр рейтинга
                        rating = item.get("rating", 0.0)
                        if rating > 0 and rating < 4.0:
                            if not is_from_cache: print("✕ (Низкий рейтинг)")
                            continue

                        # 🏆 Расчет оценки (если не из кеша или пересчитываем)
                        if not is_from_cache:
                            # 🎁 Анализ комплекта
                            if re.search(r'полный комплект|коробка|родная зарядка|ориг\w* блок', desc_lower):
                                item["kit"] = "✅ Полный"
                            else:
                                item["kit"] = "❓ Уточнить"

                            score = 0
                            if item.get("condition") == "Отличное" and item.get("screen") == "Без дефектов":
                                score += 3
                            if item.get("kit") == "✅ Полный":
                                score += 2
                            
                            jackpots = [
                                r'(nothing|phone)\s*(4\s*a|4\s*pro)\b',
                                r'(pixel|пиксель)\s*(9\s*pro|9pro)\b',
                                r'(edge 60|s60|edge 70)',
                                r'(oneplus|ванплас)\s+(12|12r|nord 4|nord 5|ace 3 pro)\b'
                            ]
                            if any(re.search(j, title_lower) for j in jackpots):
                                score += 5
                            
                            item["score"] = score
                            print(f"✅ Оценка: {score}")
                            cache[url] = item.copy()
                            new_added += 1
                        else:
                            cached_used += 1
                            print(f"  [{idx}/{len(unique_items)}] {item['title'][:20]}... (из кеша)")

                        filtered_items.append(item)
                    except Exception as e:
                        print(f"  ⚠ Ошибка при анализе {item.get('url')}: {e}")
                        continue
            
            await browser.close()
        
        # Prune cache
        current_urls = {it["url"] for it in unique_items}
        stale_keys = [k for k in list(cache.keys()) if k not in current_urls]
        for k in stale_keys:
            del cache[k]
        
        save_cache(cache)
        print(f"🗂️ Кеш обновлен: {len(cache)} объявлений (использовано {cached_used}, новых {new_added})")

        if filtered_items:
            filtered_items.sort(key=lambda x: (-x.get("score", 0), x.get("price", 0)))
            total = save_to_excel(filtered_items, OUTPUT_FILE)
            print(f"\n✨ Готово! Сохранено {total} предложений в {OUTPUT_FILE}")
        else:
            print("\n⚠ Подходящих предложений не найдено.")
    except Exception as e:
        print(f"\n❌ Критическая ошибка в работе парсера: {e}")

if __name__ == "__main__":
    asyncio.run(main())